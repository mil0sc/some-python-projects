# transpose-excel-data.py
# Transposes data from an original Excel spreadsheet to a new spreadsheet, swapping rows and columns.
# This script loads an existing Excel file, transposes its content, and saves the transposed data into a new file.

import openpyxl


def transpose_spreadsheet(original_file, transposed_file):
    """
    Transposes the content of an Excel file from rows to columns and vice versa.

    Args:
    original_file (str): Path to the original Excel file.
    transposed_file (str): Path for saving the transposed Excel file.
    """
    # Load the original workbook and select the active worksheet.
    wb = openpyxl.load_workbook(original_file)
    sheet = wb.active

    # Create a new workbook and select the active worksheet for transposed data.
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Loop through each cell in the original sheet and transpose its content.
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            # Copy the cell value to the new, transposed location in the new sheet.
            new_sheet.cell(row=j, column=i).value = sheet.cell(row=i, column=j).value

    # Save the new workbook with the transposed data.
    new_wb.save(transposed_file)


# Example usage of the function with specified original and transposed file names.
transpose_spreadsheet('original_PhoneNumbers.xlsx', 'transposed_PhoneNumbers.xlsx')

print("The spreadsheet has been transposed and saved as 'transposed_PhoneNumbers.xlsx'.")
