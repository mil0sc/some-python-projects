# format-sales-data-excel.py
# Reads sales data from an Excel file, applies conditional formatting to highlight rows based on sales targets,
# and writes the formatted data to a new Excel file. Rows where sales meet or exceed targets are highlighted in green,
# while those that don't are highlighted in red.

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Load the sales data from an Excel file into a DataFrame.
df = pd.read_excel("SalesData.xlsx")

# Initialize a new Excel workbook and select the active worksheet.
wb = Workbook()
ws = wb.active

# Populate the header row with column names from the DataFrame.
for col_num, value in enumerate(df.columns.values, start=1):
    ws.cell(row=1, column=col_num, value=value)

# Iterate over the DataFrame rows to write data to the workbook.
# Apply conditional formatting based on sales target achievement.
for row_num, row_data in enumerate(df.values, start=2):
    for col_num, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=value)

        # Conditional formatting: Green fill if Units Sold meets/exceeds Target, Red otherwise.
        # Adjust column indexes based on actual data layout if necessary.
        if row_data[1] >= row_data[2]:  # Assuming 'Units Sold' is the second column and 'Target' is the third.
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Save the workbook with the formatted sales data to a new file.
wb.save("FormattedSalesData.xlsx")

print("Sales data has been formatted and saved to 'FormattedSalesData.xlsx'.")
