# spreadsheet-to-text-converter.py
# Extracts data from each column of an Excel spreadsheet and saves it into individual text files.
# Each column's data is saved into a separate text file, named after the column letter with a '.txt' extension.

import openpyxl


def spreadsheet_to_text_files(spreadsheet_name):
    """
    Converts each column in an Excel spreadsheet into a separate text file.

    Args:
    spreadsheet_name (str): The name of the Excel file to convert.
    """
    # Load the workbook and select the active worksheet.
    wb = openpyxl.load_workbook(spreadsheet_name)
    ws = wb.active

    # Determine the maximum number of columns and rows in the worksheet.
    max_column = ws.max_column
    max_row = ws.max_row

    # Iterate over each column in the worksheet.
    for col in range(1, max_column + 1):
        # Create a new text file for the current column, naming it after the column letter.
        file_name = f'Column_{chr(64 + col)}.txt'
        with open(file_name, 'w') as file:
            # Write the value of each cell in the current column to the text file.
            for row in range(1, max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                # Ensure only non-empty cells are written to the file.
                if cell_value is not None:
                    file.write(str(cell_value) + '\n')

    print(f"Text files have been created from the spreadsheet '{spreadsheet_name}'.")


# Example usage:
# Specify the name of the Excel spreadsheet to convert.
spreadsheet_name = 'text_contents.xlsx'
# Execute the function to convert the spreadsheet into text files.
spreadsheet_to_text_files(spreadsheet_name)
