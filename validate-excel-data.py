# validate-excel-data.py
# Scans through an Excel workbook to identify and report blank rows and cells containing invalid data.
# It defines valid data as non-negative numbers and non-blank cells, but this criteria can be adjusted as needed.

import openpyxl


def is_valid_data(value):
    """
    Determines if the provided cell value is considered valid.

    Args:
    value: The value to check for validity.

    Returns:
    bool: True if the value is valid, False otherwise.
    """
    # Checks for blank cells or negative numeric values.
    if value is None or (isinstance(value, (int, float)) and value < 0):
        return False
    return True


print('Opening workbook...')
# Load the workbook and select the active worksheet.
wb = openpyxl.load_workbook('.xlsx')
sheet = wb.active

# Lists to store the locations of blank rows and invalid data cells.
blank_rows = []
invalid_data_cells = []

# Iterate through each row and column in the worksheet.
for row in range(1, sheet.max_row + 1):
    is_row_blank = True  # Assume the row is blank until proven otherwise.
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=row, column=col).value
        # Check if the cell disrupts the assumption of a blank row.
        if cell_value:
            is_row_blank = False
        # Check for invalid data according to the defined criteria.
        if not is_valid_data(cell_value):
            invalid_data_cells.append((row, col))
    # If the row is confirmed to be blank, add it to the list.
    if is_row_blank:
        blank_rows.append(row)

# Report findings on blank rows and invalid data cells.
if blank_rows:
    print(f"Found blank rows at: {', '.join(map(str, blank_rows))}")
if invalid_data_cells:
    for cell in invalid_data_cells:
        print(f"Invalid data found at cell {sheet.cell(row=cell[0], column=cell[1]).coordinate}")

# If no issues are found, inform the user.
if not blank_rows and not invalid_data_cells:
    print("No blank rows or invalid data found!")
