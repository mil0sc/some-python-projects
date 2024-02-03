# compare-monthly-sales-excel.py
# Reads sales data for January and February from an Excel workbook, compares the sales of items between the two months,
# and prints out the differences in sales volumes, including items sold exclusively in one month.

import openpyxl

print('Opening workbook...')
# Load the workbook and specify the sheets for January and February sales data.
wb = openpyxl.load_workbook('.xlsx')
sheet1 = wb['January']
sheet2 = wb['February']

# Initialize dictionaries to hold sales data for each month.
january_sales = {}
february_sales = {}

# Populate the january_sales dictionary with item names as keys and sold quantities as values.
for row in range(2, sheet1.max_row + 1):
    item = sheet1.cell(row=row, column=1).value
    sold = float(sheet1.cell(row=row, column=2).value)
    january_sales[item] = sold

# Populate the february_sales dictionary similarly for February.
for row in range(2, sheet2.max_row + 1):
    item = sheet2.cell(row=row, column=1).value
    sold = float(sheet2.cell(row=row, column=2).value)
    february_sales[item] = sold

# Compare sales volumes between January and February for each item and print the differences.
for item, sold in january_sales.items():
    if item in february_sales:
        difference = sold - february_sales[item]
        if difference > 0:
            print(f"{item} had {difference} more sales in January than in February.")
        elif difference < 0:
            print(f"{item} had {-difference} more sales in February than in January.")
        else:
            print(f"{item} had the same number of sales in both months.")
    else:
        print(f"{item} was only sold in January.")

# Additionally, identify items that were only sold in February.
for item in february_sales:
    if item not in january_sales:
        print(f"{item} was only sold in February.")
