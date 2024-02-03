# convert-excel-to-csv.py
# Converts each sheet in Excel files within a specified directory to individual CSV files,
# naming the CSV files based on the original Excel file and sheet names.

import os
import csv
from openpyxl import load_workbook


def convert_excel_sheets_to_csv(excel_files_dir):
    """
    Converts all sheets in each Excel file in the specified directory to CSV files.

    Args:
    excel_files_dir (str): The directory containing the Excel files to be converted.
    """
    # Change the working directory to where the Excel files are located.
    os.chdir(excel_files_dir)

    # Iterate over all files in the specified directory.
    for excel_file in os.listdir('.'):
        # Process only files with the .xlsx extension.
        if not excel_file.endswith('.xlsx'):
            continue

        # Load the current Excel workbook.
        wb = load_workbook(excel_file)

        # Convert each sheet in the workbook to a CSV file.
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            csv_filename = f"{os.path.splitext(excel_file)[0]}_{sheet_name}.csv"

            # Open a new CSV file and set up a writer.
            with open(csv_filename, 'w', newline='') as csv_file:
                csv_writer = csv.writer(csv_file)

                # Write each row in the sheet to the CSV file.
                for row in sheet.iter_rows(values_only=True):
                    csv_writer.writerow(row)

        # Ensure the workbook is closed to free resources.
        wb.close()

    print('All Excel files have been converted to CSV files.')


# Example usage of the function with the path to the directory containing Excel files.
excel_files_dir = 'path_to_excel_files'  # Replace with the actual path to your Excel files.
convert_excel_sheets_to_csv(excel_files_dir)
