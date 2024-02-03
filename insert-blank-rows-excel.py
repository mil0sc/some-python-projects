# insert-blank-rows-excel.py
# This program inserts a specified number of blank rows into an Excel spreadsheet starting from a specified row.
# It takes three command-line arguments: the row number to start at, the number of blank rows to insert, and the filename of the Excel spreadsheet.

import sys
from openpyxl import load_workbook

# Validate command-line arguments and provide usage instructions if incorrect.
if len(sys.argv) != 4:
    print("Usage: python insert-blank-rows-excel.py <N> <M> <filename>")
    sys.exit()

# Assign command-line arguments to variables for ease of use.
N = int(sys.argv[1])  # Row number from where to start inserting blank rows.
M = int(sys.argv[2])  # Number of blank rows to insert.
filename = sys.argv[3]  # Excel file to modify.

# Load the workbook and select the active sheet for reading data.
wb = load_workbook(filename)
ws = wb.active

# Create a new workbook and clear any existing sheets to start fresh.
wb_new = load_workbook(filename)
for sheet in wb_new.sheetnames:
    wb_new.remove(wb_new[sheet])
ws_new = wb_new.create_sheet(title="Sheet")  # Create a new sheet in the workbook.

# Insert blank rows by adjusting the row index for each cell after the specified row.
for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
    # Determine new row index, adjusting for inserted blank rows as needed.
    new_row_idx = row[0].row + M if row[0].row >= N else row[0].row

    # Copy each cell to the new worksheet, preserving content and format.
    for cell in row:
        ws_new.cell(row=new_row_idx, column=cell.column, value=cell.value)

# Save the modified spreadsheet. Consider saving as a new file to avoid overwriting.
wb_new.save(filename)
print(f"Spreadsheet modified with {M} blank rows inserted starting from row {N}.")
