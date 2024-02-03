# generate-custom-greetings.py
# Reads user preferences for favorite colors from an Excel file named 'UserPreferences.xlsx',
# and generates personalized greetings for each user with a color-themed dashboard suggestion.

import openpyxl

# Load the workbook and select the active worksheet to access the data.
wb = openpyxl.load_workbook('UserPreferences.xlsx')
sheet = wb.active

# Initialize a dictionary to store user preferences with usernames as keys and favorite colors as values.
user_preferences = {}

# Iterate over the rows in the spreadsheet starting from the second row to skip the header.
for row in range(2, sheet.max_row + 1):
    # Extract the user's name and their favorite color from each row.
    user_name = sheet.cell(row=row, column=1).value
    user_color = sheet.cell(row=row, column=2).value
    # Store the extracted data in the user_preferences dictionary.
    user_preferences[user_name] = user_color

# Generate and print a personalized greeting for each user based on their favorite color.
for user, color in user_preferences.items():
    # Construct a greeting message incorporating the user's name and color preference.
    greeting = f"Hello {user}, here's a {color.lower()}-themed dashboard for you!"
    print(greeting)
