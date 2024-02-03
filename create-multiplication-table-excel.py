# create-multiplication-table-excel.py
# Generates an Excel spreadsheet with a multiplication table up to a specified number.
# It accepts a single command-line argument representing the size of the multiplication table to generate.

import sys
from openpyxl import Workbook
from openpyxl.styles import Font

# Ensure correct usage with a single command-line argument for the table size.
if len(sys.argv) != 2:
    print("Usage: py create-multiplication-table-excel.py <number>")
    sys.exit()

# Validate the provided number and convert it to an integer.
try:
    N = int(sys.argv[1])
except ValueError:
    print("Please provide a valid integer.")
    sys.exit()

# Initialize a new Excel workbook and select the active worksheet.
wb = Workbook()
ws = wb.active

# Fill the first row and column with numbers 1 through N to label the table.
for i in range(1, N+1):
    ws.cell(row=i+1, column=1).value = i  # Label the rows.
    ws.cell(row=1, column=i+1).value = i  # Label the columns.

# Populate the table with multiplication results.
for row in range(1, N+1):
    for col in range(1, N+1):
        ws.cell(row=row+1, column=col+1).value = row * col

# Apply bold formatting to the labels for clarity.
bold_font = Font(bold=True)
for i in range(1, N+2):
    ws.cell(row=1, column=i).font = bold_font  # Make column labels bold.
    ws.cell(row=i, column=1).font = bold_font  # Make row labels bold.

# Save the filled workbook as an Excel file.
file_name = "multiplication_table.xlsx"
wb.save(file_name)
print(f"Multiplication table saved as {file_name}")
