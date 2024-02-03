# analyze-discounts-excel.py
# Opens an Excel workbook and iterates through rows of a specified sheet named 'znizki' to identify items
# with discounts greater than 10% of their original price. It prints details of such items, including
# the item name, original price, and discounted price.

import openpyxl

print('Opening workbook...')
# Load the workbook and access the specified sheet.
wb = openpyxl.load_workbook('.xlsx')
sheet = wb['znizki']

# Iterate through each row in the sheet, starting from row 2 to skip the header.
for row in range(2, sheet.max_row + 1):
    # Extract item name, original price, and discounted price from the current row.
    item = sheet.cell(row=row, column=1).value
    ogCena = float(sheet.cell(row=row, column=2).value)
    przecena = float(sheet.cell(row=row, column=3).value)

    # Calculate the discount amount.
    discount = ogCena - przecena

    # Determine if the discount is greater than 10% of the original price.
    if discount >= 0.1 * ogCena:
        # Print item details if the discount criteria is met.
        print(f'{item} has a discount greater than 10%. Original Price: ${ogCena}, Discounted Price: ${przecena}')
